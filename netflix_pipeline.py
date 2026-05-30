#importing all required libraries 
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains as ac
from selenium.webdriver.common.keys import Keys

import csv,pickle,unicodedata
import openpyxl as xl,re,time

import os,csv

#delete files for adding data instead of updating data
file_path = "series.csv"
if os.path.exists(file_path):
    os.remove(file_path)
open(file_path, "w", encoding="utf-8").close() 
file_path = "films.csv"
if os.path.exists(file_path):
    os.remove(file_path)
open(file_path, "w", encoding="utf-8").close() 
file_path = "netflix_backlog.xlsx"
if os.path.exists(file_path):
    os.remove(file_path)

try :
    #run browser
    browser = webdriver.Firefox()
    browser.get("https://www.netflix.com")
    wait = WebDriverWait(browser,10)
    action = ac(browser)
    #load cookies
    with open("netflix_cookies.pkl","rb") as f:
        cookies = pickle.load(f)
    for cookie in cookies :
        if "sameSite" in cookie :
            cookie.pop("sameSite")
        browser.add_cookie(cookie)
    browser.refresh()
    #create a dictionary with profile name as key and profile element as value . later using the name select profile of your choice
    profiles = WebDriverWait(browser,60).until(ec.presence_of_all_elements_located((By.CSS_SELECTOR,'li.profile')))
    profileDict = {}
    for profile in profiles:
        nameelement = profile.find_element(By.CLASS_NAME,"profile-name")
        name  = nameelement.text
        profileDict[name] = profile
    profile = profileDict["M.J"]
    profile.click()

    alltitles = []
    for j in range(4):
        time.sleep(4)
        #find the contents in "continue watching" that are films and series , few of those contents are not accessible at this point so filter which are only accessible
        contents = wait.until(ec.presence_of_all_elements_located((By.CSS_SELECTOR,"#row-2 .slider-item")))
        filtered_content = [content for content in contents if re.search(r"slider-item-\d+",content.get_attribute("class"))]

        for content in filtered_content:
            #find title and add it to list of alltitles
            title_elem = content.find_element(By.CSS_SELECTOR,"div.title-card > div > a.slider-refocus")
            title = title_elem.get_attribute("aria-label")
            if title not in alltitles :
                alltitles.append(title)
            else:
                continue
            print("scraping data of ",title)
            #find the title card element
            card = WebDriverWait(content,10).until(ec.presence_of_element_located((By.CSS_SELECTOR,"div.title-card")))
            time.sleep(1)
            #move the element into the middle area of the screen
            browser.execute_script("arguments[0].scrollIntoView({block:'center'});",card)
            time.sleep(1)
            #move the cursor to the center of the card and then ofset it to make the mouse movement natural for the website 
            action.move_to_element(card).pause(1).move_by_offset(1, 1).perform()
            time.sleep(1)
            #find the expand button and click it
            more = WebDriverWait(browser,10).until(ec.presence_of_element_located((By.CSS_SELECTOR,'div.buttonControls--expand-button > button')))
            browser.execute_script("arguments[0].click();",more)
            #wait until it loads and to know whether you selected a film or series search for element with class name episode selector 
            wait.until(ec.presence_of_element_located((By.CLASS_NAME,"detail-modal-container")))
            episodeSelector = browser.find_elements(By.CLASS_NAME,"episodeSelector")
            #the episode selector variable contains a list of elments if it is a film the list will be empty else it will have a element
            if len(episodeSelector)>0:
                #check for a season label if it is present then the series has multiple seasons else only 1 season
                try:
                    current_elem = wait.until(ec.presence_of_all_elements_located((By.CLASS_NAME,"allEpisodeSelector-season-label")))
                except Exception:
                    current_elem = []
                #if a series have multiple seasons then find the text like this "Season 3"
                if len(current_elem) > 0 :
                    current_season = wait.until(ec.presence_of_element_located((By.CLASS_NAME,"allEpisodeSelector-season-label"))).text[0:-1]
                else :
                    current_season = "1"
                #season number that is present in the string "Season 3"
                numbers_in_string = re.findall(r"\d+",current_season)
                #if number is present fine if not that means the number must be in roman number or anyother format
                if len(numbers_in_string) > 0:
                    current_season_num = int(numbers_in_string[0])
                else :
                    #convert the non ASCII character into number
                    current_season_num = int(unicodedata.numeric(current_season[-1])) if current_season[-1].isnumeric() else -1
                
                seasons = {}
                #find dropdown box if it exists select the dropdown box and click it 
                dropdown_elem = browser.find_elements(By.CSS_SELECTOR,"div.episodeSelector-dropdown>div>button")
                if len(dropdown_elem) > 0:
                    dropbox = wait.until(ec.presence_of_element_located((By.CSS_SELECTOR,"div.episodeSelector-dropdown>div>button")))
                    dropbox.click()
                    #select elements that are present in the dropdown box which are the seasons 
                    dropbox_elements = wait.until(ec.presence_of_all_elements_located((By.CSS_SELECTOR,"div.episodeSelector-dropdown>div ul>li>div")))
                    for element in dropbox_elements:
                        #format is "Season 1 (12 episodes)" so seperate the string 
                        season , episodes = element.text.split("\n")
                        #if season strung contain non ASCII digit convert it
                        if not re.search(r'\d+', season):
                            season = season.strip()
                            season = season[:-1]+ str(int(unicodedata.numeric(season[-1]))) if season[-1].isnumeric() else ' 1'
                        #seperate the string again to find number of episodes 
                        num , string = episodes.split(" ")
                        seasons[season.strip()[-1]] = int(num[1:])
                else :
                    seasons["1"] = -1
                #find how much time you watched each episode
                progress_elements = wait.until(ec.presence_of_all_elements_located((By.CLASS_NAME,"titleCard-progress")))
                watched_episodes = []
                #add all viewed episodes to a list
                for progress_element in progress_elements:
                    progress = float(progress_element.get_attribute("value"))
                    watched_episodes.append(progress)
                #remove video from the list if only half video or less is watched
                for value in watched_episodes:
                    if value < 0.5 :
                        watched_episodes.remove(value)
                #assuming your all the previous seasons are watched create a empty list to track remaining episodes
                episodes_left = {}
                i = current_season_num 
                data = []
                while(i<=len(seasons)):
                    #in the current season remove watched episodes from total episodes
                    if i == current_season_num:
                        episodes_left[f"Season{i}"] = seasons[str(i)] - len(watched_episodes)
                        data.append([title,f"Season {i}",episodes_left[f"Season{i}"]])
                    #in later season add all episodes to dictionary
                    else:
                        episodes_left[f"Season{i}"] = seasons[str(i)] 
                        data.append([title,f"Season {i}",f"{episodes_left[f'Season{i}']} left"])
                    i +=1
                #save the data in series.csv file
                with open("series.csv","a",newline="",encoding="utf-8") as f:
                    csvwriter = csv.writer(f)
                    csvwriter.writerows(data)
            #the content you are looking at is a film 
            else :
                #find the progress bar element
                summary = wait.until(ec.presence_of_element_located((By.CLASS_NAME,"summary"))).text
                listofnumbers = re.findall(r"\d+",summary)
                #netflix follow 2 formats to represent progress "23 min of 56" or "35 min left"
                #if the list contains 2 numbers then time left is 2nd element - 1st
                if len(listofnumbers) > 1:
                    time_left = f"{int(listofnumbers[1]) - int(listofnumbers[0])}minutes Left"
                #if only one element is present that is the time left 
                else :
                    time_left = f"{int(listofnumbers[0])}minutes Left"
                data = [title,time_left]
                #save the data in films.csv file
                with open("films.csv","a",newline="",encoding="utf-8") as f:
                    csvwriter = csv.writer(f)
                    csvwriter.writerow(data)
            #one content is saved now press escape to move on to next element
            browser.find_element(By.TAG_NAME,"body").send_keys(Keys.ESCAPE)

            
        time.sleep(2)
        #after all the elements that are visible are saved next we need to move to right
        #find the right handle and press it 
        handles = wait.until(ec.presence_of_all_elements_located((By.CSS_SELECTOR, "#row-2 span.handle>b")))
        right_arrow_area = handles[-1]
        time.sleep(2)
        #make the handle be present in the center and click it 
        browser.execute_script("arguments[0].scrollIntoView({block:'center',inline:'center'});",right_arrow_area)
        browser.execute_script("arguments[0].click();",right_arrow_area)
        time.sleep(2)


except Exception :
    print("an exception has occured it is %s"%(Exception))

finally: 
    browser.quit()

#create a excel sheet and add contents
wb = xl.Workbook()
sheet = wb.active
sheet.title = "MJ's continue watching"
sheet["A1"] = "TITLE"
sheet["B1"] = "Episodes or minutes left"
sheet.merge_cells("B1:C1")

with open("films.csv","r",encoding="utf-8") as file :
    rows = csv.reader(file)
    for j,row in enumerate(rows):
        movie_title = row[0]
        minutes_left = row[1]
        sheet["A"+str(j+2)] = movie_title
        sheet["B"+str(j+2)] = minutes_left
        sheet.merge_cells(f"{'B'+str(j+2)}:{'C'+str(j+2)}")
    j+=2
with open("series.csv","r",encoding="utf-8") as file :
    rows = csv.reader(file)
    for row in rows:
        
        sheet["A"+str(j+1)] = row[0]
        sheet["B"+str(j+1)] = row[1]
        sheet["C"+str(j+1)] = row[2]
        j+=1
sheet.column_dimensions["A"].width = 30
sheet.column_dimensions["C"].width = 30
sheet.column_dimensions["B"].width = 30
sheet.freeze_panes = "D2"
wb.save("netflix_backlog.xlsx")
